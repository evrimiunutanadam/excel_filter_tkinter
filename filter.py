import openpyxl
from openpyxl import Workbook
import tkinter as tk




def clicked():

    print("deneme")
    
    try:
        row_check = ""
        filter_length = int(filter_length_txt.get())
        filter_column_indice = int(filter_no_txt.get())
        test_indice = int(search_no_txt.get())
        file_location = str(source_adress_txt.get())
        wb = openpyxl.load_workbook(file_location)

        sheet = wb.active

        clean_rows = []

        spam_rows = []

        if filter_length > sheet.max_row:
            filter_length=sheet.max_row


        for i in range(2, sheet.max_row + 1):
            control = 0


            for j in range(2, filter_length + 1):
                tested_cell = sheet.cell(row=i, column=test_indice)
                filter_cell = sheet.cell(row=j, column=filter_column_indice)
                searched_word = str(tested_cell.value)
                filter_word = str(filter_cell.value)
                if filter_word in searched_word:
                    onerow = []


                    for k in range(1, sheet.max_column + 1):
                        onerow.append(sheet.cell(row=tested_cell.row, column=k).value)

                    spam_rows.append(onerow)
                    control = 1
                    break

            if control == 0:
                onerow = []
                for k in range(1, sheet.max_column + 1):
                    onerow.append(sheet.cell(row=tested_cell.row, column=k).value)
                row_check = row_check + "row: {} clear: - / {}\n".format(tested_cell.row, searched_word)

                clean_rows.append(onerow)
            else:
                row_check = row_check + "row: {}  Detected Filter: {} for value {} (detected)\n".format(tested_cell.row,
                                                                                                    filter_word,
                                                                                                    searched_word)
        T.insert(tk.END, row_check)


        cleared_workbook = Workbook()
        ws = cleared_workbook.active
        ws.title = "cleared words"

        g = 1
        k = 1

        for items in clean_rows:
            for values in items:
                ws.cell(row=g, column=k).value = values
                k = k + 1
            g = g + 1
            k = 1
        new_excel_name = new_excel_name_txt.get()

        cleared_workbook.save(new_excel_name)
        verification_label = tk.Label(window, text="Filtered data succesfully saved as,{}".format(new_excel_name))
        verification_label.grid(column=0, row=9)

        window.update()

    except:
        error_label = tk.Label(window,"Şu sebeplerden dolayı bir hata oluştu:\n -Yeni dosya ismine .xlsx uzantısı eklenmedi.\n-Küçük kutulara sayı dışında bir karakter girildi \n-Bütün kutular doldurulmadan program çalıştırıldı.\n-Okunacak kaynak dosyanın pc adresi yanlış verildi. ")
        error_label.grid(column=0, row=9)







window = tk.Tk()

window.title("Excel Filter")

window.geometry('700x800')


lbl = tk.Label(window, text="Source Excel File Adress")
lbl.grid(column=0, row=0)

lbl4 = tk.Label(window, text="Name Of The New File:")
lbl4.grid(column=0,row=7)

lbl = tk.Label(window, text="Filter Column No:")
lbl.grid(column=0, row=2)

lbl2 = tk.Label(window, text="Search Column No:")
lbl2.grid(column=0,row=3)

lbl3 = tk.Label(window, text="Filter Column Length:")
lbl3.grid(column=0,row=4)

lbl_instructions=tk.Label(window,text="Instructions\n\n1-Write path of \nthe excel file\nthat you want to\n filter\n\n2-Write number of \nfilter column, \nsearch column \nand length \nof filter \n column without \nany letters.\n\n3-Write name of the\n new excel file \nwith extension \n.xlsx\n That will contain \nclear data.\n ")
lbl_instructions.grid(column=1,row=6)


source_adress_txt=tk.Entry(window,width=50)
filter_no_txt= tk.Entry(window, width=6)
search_no_txt=tk.Entry(window,width=6)
filter_length_txt=tk.Entry(window,width=6)
new_excel_name_txt=tk.Entry(window,width=50)

source_adress_txt.grid(column=0,row=1)
new_excel_name_txt.grid(column=0,row=8)
filter_no_txt.grid(column=1, row=2)
search_no_txt.grid(column=1,row=3)
filter_length_txt.grid(column=1,row=4)


btn = tk.Button(window, text="start",command=clicked)

btn.grid(column=0, row=5)
btn.config(height=2,width=40)

T=tk.Text(window,height=30,width=70,bg="lightgoldenrodyellow")
S=tk.Scrollbar(window)

S.config(command=T.yview)
T.config(yscrollcommand=S.set)

T.grid(column=0,row=6)

window.mainloop()